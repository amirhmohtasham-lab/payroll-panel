"""Shared helpers for workbook auditing: cell scanning, Persian digit normalization, highlighting."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

_PERSIAN_DIGITS = "۰۱۲۳۴۵۶۷۸۹"
_ARABIC_DIGITS = "٠١٢٣٤٥٦٧٨٩"


def normalize_digits(value: str) -> str:
    table = str.maketrans(
        _PERSIAN_DIGITS + _ARABIC_DIGITS,
        "0123456789" + "0123456789",
    )
    return value.translate(table)


def to_number(value: object) -> float | None:
    """Best-effort conversion of a cell value (possibly Persian-digit, comma-separated) to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_digits(str(value)).strip()
    text = text.replace(",", "").replace("،", "").replace(" ", "")
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_header_cell(ws: Worksheet, keywords: list[str], *, max_row: int = 15) -> tuple[int, int] | None:
    """Search the first `max_row` rows for a cell whose text contains any of `keywords`."""
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, ws.max_row)):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            if any(k in text for k in keywords):
                return cell.row, cell.column
    return None


@dataclass
class Issue:
    severity: str  # "error" | "warn"
    code: str
    sheet: str
    message: str


def open_workbook_data_only(path: Path) -> Workbook:
    return load_workbook(path, data_only=True)


def open_workbook_for_write(path: Path) -> Workbook:
    return load_workbook(path)


def save_highlighted(wb: Workbook, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
