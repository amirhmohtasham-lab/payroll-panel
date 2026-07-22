"""Upload + audit flow: validate xlsx, hash, audit, persist, duplicate detection."""
from __future__ import annotations

import io

from openpyxl import Workbook


def _payroll_xlsx_bytes(*, worker_amount: int = 1000) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "سرکارگر"
    ws["B1"] = "علی محمدی"
    ws["A4"] = "نام کارگر"
    ws["B4"] = "جمع دریافتی"
    ws["A5"] = "رضا رضایی"
    ws["B5"] = worker_amount
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _login_operator(client, operator_user):
    res = client.post("/api/login", json={"username": "operator1", "password": "op-pass-123"})
    assert res.status_code == 200


def test_upload_rejects_non_xlsx(client, operator_user):
    _login_operator(client, operator_user)
    files = {"file": ("payroll.txt", b"not an excel file", "text/plain")}
    data = {"month_key": "1405-01", "month_label": "فروردین"}
    res = client.post("/api/upload", files=files, data=data)
    assert res.status_code == 400


def test_upload_requires_auth(client):
    files = {"file": ("payroll.xlsx", _payroll_xlsx_bytes(), "application/vnd.openxmlformats")}
    data = {"month_key": "1405-01", "month_label": "فروردین"}
    res = client.post("/api/upload", files=files, data=data)
    assert res.status_code == 401


def test_upload_success_and_persistence(client, operator_user, accountant_user):
    _login_operator(client, operator_user)
    files = {"file": ("payroll.xlsx", _payroll_xlsx_bytes(), "application/vnd.openxmlformats")}
    data = {"month_key": "1405-01", "month_label": "فروردین ۱۴۰۵"}
    res = client.post("/api/upload", files=files, data=data)
    assert res.status_code == 200
    body = res.json()
    assert body["ok"] is True
    assert body["error_count"] == 0
    assert body["record"]["month_key"] == "1405-01"

    client.post("/api/logout")
    client.post("/api/login", json={"username": "admin1", "password": "acc-pass-123"})
    months_res = client.get("/api/months")
    assert months_res.status_code == 200
    assert months_res.json()["summary"]["month_count"] == 1


def test_upload_duplicate_month_without_replace(client, operator_user):
    _login_operator(client, operator_user)
    files = {"file": ("payroll.xlsx", _payroll_xlsx_bytes(), "application/vnd.openxmlformats")}
    data = {"month_key": "1405-02", "month_label": "اردیبهشت"}
    res1 = client.post("/api/upload", files=files, data=data)
    assert res1.status_code == 200

    files2 = {"file": ("payroll2.xlsx", _payroll_xlsx_bytes(worker_amount=2000), "application/vnd.openxmlformats")}
    res2 = client.post("/api/upload", files=files2, data=data)
    assert res2.status_code == 409
    assert res2.json()["duplicate"] is True


def test_upload_duplicate_month_with_replace(client, operator_user):
    _login_operator(client, operator_user)
    files = {"file": ("payroll.xlsx", _payroll_xlsx_bytes(), "application/vnd.openxmlformats")}
    data = {"month_key": "1405-03", "month_label": "خرداد"}
    res1 = client.post("/api/upload", files=files, data=data)
    assert res1.status_code == 200

    files2 = {"file": ("payroll2.xlsx", _payroll_xlsx_bytes(worker_amount=2000), "application/vnd.openxmlformats")}
    data2 = {"month_key": "1405-03", "month_label": "خرداد", "replace": "true"}
    res2 = client.post("/api/upload", files=files2, data=data2)
    assert res2.status_code == 200
    assert res2.json()["record"]["original_filename"] == "payroll2.xlsx"


def test_upload_invalid_month_key(client, operator_user):
    _login_operator(client, operator_user)
    files = {"file": ("payroll.xlsx", _payroll_xlsx_bytes(), "application/vnd.openxmlformats")}
    data = {"month_key": "not-a-month", "month_label": "x"}
    res = client.post("/api/upload", files=files, data=data)
    assert res.status_code == 400
