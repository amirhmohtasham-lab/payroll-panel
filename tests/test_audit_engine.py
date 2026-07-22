"""Unit tests for the audit_engine package (payroll + fertilizer workbook auditing)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from audit_engine import fertilizer as fertilizer_engine
from audit_engine import payroll as payroll_engine


def _make_payroll_workbook(path: Path, *, negative_amount: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "سرکارگر"
    ws["B1"] = "علی محمدی"
    ws["A2"] = "چاه"
    ws["B2"] = "چاه ۱"
    ws["A4"] = "نام کارگر"
    ws["B4"] = "جمع دریافتی"
    ws["A5"] = "رضا رضایی"
    ws["B5"] = -100 if negative_amount else 1000
    ws["A6"] = "حسن حسنی"
    ws["B6"] = 2000
    wb.save(path)


def test_payroll_audit_happy_path(tmp_path: Path):
    wb_path = tmp_path / "payroll.xlsx"
    _make_payroll_workbook(wb_path)

    result = payroll_engine.audit_workbook(wb_path)

    assert len(result.sheets) == 1
    sheet = result.sheets[0]
    assert sheet.worker_rows == 2
    assert sheet.foreman == "علی محمدی"
    assert sheet.workplace == "چاه ۱"
    assert sheet.worker_gross == 3000
    assert not any(i.severity == "error" for i in result.issues)


def test_payroll_audit_flags_negative_amount(tmp_path: Path):
    wb_path = tmp_path / "payroll_negative.xlsx"
    _make_payroll_workbook(wb_path, negative_amount=True)

    result = payroll_engine.audit_workbook(wb_path)

    error_codes = [i.code for i in result.issues if i.severity == "error"]
    assert "NEGATIVE_AMOUNT" in error_codes


def test_payroll_audit_missing_header(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "شماره سریال"
    ws["B1"] = "توضیحات"
    wb_path = tmp_path / "no_header.xlsx"
    wb.save(wb_path)

    result = payroll_engine.audit_workbook(wb_path)
    assert any(i.code == "NO_HEADER" for i in result.issues)


def test_payroll_write_highlighted_workbook(tmp_path: Path):
    wb_path = tmp_path / "payroll.xlsx"
    _make_payroll_workbook(wb_path, negative_amount=True)
    result = payroll_engine.audit_workbook(wb_path)

    dest = tmp_path / "highlighted.xlsx"
    payroll_engine.write_highlighted_workbook(result, wb_path, dest)

    assert dest.exists()


def _make_fertilizer_workbook(path: Path, *, negative_quantity: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "نوع کود"
    ws["B1"] = "مقدار مصرف"
    ws["A2"] = "اوره"
    ws["B2"] = -50 if negative_quantity else 100
    ws["A3"] = "فسفات"
    ws["B3"] = 200
    wb.save(path)


def test_fertilizer_audit_happy_path(tmp_path: Path):
    wb_path = tmp_path / "fertilizer.xlsx"
    _make_fertilizer_workbook(wb_path)

    result = fertilizer_engine.audit_workbook(wb_path)

    assert len(result.rows) == 2
    assert result.fertilizers == {"اوره", "فسفات"}
    assert not any(i.severity == "error" for i in result.issues)


def test_fertilizer_audit_flags_negative_quantity(tmp_path: Path):
    wb_path = tmp_path / "fertilizer_negative.xlsx"
    _make_fertilizer_workbook(wb_path, negative_quantity=True)

    result = fertilizer_engine.audit_workbook(wb_path)

    error_codes = [i.code for i in result.issues if i.severity == "error"]
    assert "NEGATIVE_QUANTITY" in error_codes


def test_to_number_handles_persian_digits():
    from audit_engine.common import to_number

    assert to_number("۱۲۳۴") == 1234
    assert to_number("1,234") == 1234
    assert to_number("-") is None
    assert to_number(None) is None
    assert to_number(42) == 42.0
