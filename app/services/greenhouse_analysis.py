"""Greenhouse climate analysis core — updated for the payroll panel backend.

Based on the updated Hydroponic-Greenhouse-Data-Analysis script:
  - all setpoints are configurable (RH, temperature day/night, VPD target, day/night hours, irrigation window)
  - dynamic temperature breach levels based on setpoints
  - department colors for charts (A-D)
Supports both UTF-16LE controller CSV exports and XLSX inputs.
"""

from __future__ import annotations

import datetime
import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

DEPARTMENTS = ["A", "B", "C", "D"]
HUMIDITY_NAMES = ["Time", "Date"] + [f"HUMI {d}{n}" for d in DEPARTMENTS for n in (1, 2)]
TEMPERATURE_NAMES = ["Time", "Date"] + [f"TEMP {d}{n}" for d in DEPARTMENTS for n in range(1, 5)]
TEMP_SENSORS = [f"{d}{n}" for d in DEPARTMENTS for n in range(1, 5)]
DEPARTMENT_COLORS = {"A": "#1f77b4", "B": "#ff7f0e", "C": "#2ca02c", "D": "#d62728"}
STATISTIC_COLORS = {"Min": "#2563eb", "Avg": "#f59e0b", "Max": "#dc2626"}


def require_columns(frame: pd.DataFrame, required: list[str], label: str) -> None:
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise ValueError(f"{label} is missing: {', '.join(missing)}")


def parse_datetime(date: pd.Series, time: pd.Series, label: str) -> pd.Series:
    date_text = date.astype(str).str.strip()
    time_text = time.astype(str).str.strip().str.replace(r"^(\d{1,2}:\d{2})$", r"\1:00", regex=True)
    result = pd.to_datetime(date_text + " " + time_text, errors="coerce")
    if result.isna().any():
        rows = (np.flatnonzero(result.isna())[:20] + 2).tolist()
        raise ValueError(f"{label} has unparseable Date/Time values in spreadsheet row(s): {rows}")
    return result


def write_single_sheet(frame: pd.DataFrame, path: Path, sheet_name: str = "Sheet1") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_excel(path, index=False, sheet_name=sheet_name)


def read_greenhouse_csv(path: Path, names: list[str], label: str) -> pd.DataFrame:
    """Read the controller's UTF-16LE, tab-delimited export."""
    frame = pd.read_csv(
        path,
        sep="\t",
        encoding="utf-16le",
        skiprows=1,
        header=None,
        names=names,
        usecols=range(len(names)),
        dtype={"Time": "string", "Date": "string"},
    )
    if frame.shape[1] != len(names) or frame.empty:
        raise ValueError(f"{label}: expected {len(names)} non-empty columns; found shape {frame.shape}.")
    return frame


def read_greenhouse_source(path: Path, names: list[str], label: str) -> pd.DataFrame:
    """Read a greenhouse export: UTF-16LE CSV (controller) or XLSX."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_greenhouse_csv(path, names, label)
    frame = pd.read_excel(path)
    frame.columns = [re.sub(r"^(TEMP|HUMI)\s+", "", str(c).strip()) for c in frame.columns]
    expected = [n.split(" ", 1)[1] if " " in n else n for n in names[2:]]
    expected_full = ["Time", "Date"] + expected
    require_columns(frame, expected_full, label)
    frame = frame[expected_full].copy()
    frame.columns = names
    return frame


def read_temperature_source(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return read_greenhouse_csv(path, TEMPERATURE_NAMES, "Temperature CSV")
    return pd.read_excel(path)


def save_plot(figure, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    figure.write_html(path, include_plotlyjs="cdn")


def get_plotly():
    try:
        import plotly.express as px
        import plotly.graph_objects as go
        return px, go
    except ImportError:
        warnings.warn("plotly is not installed; interactive HTML plots were skipped.")
        return None, None


def add_period_columns(frame: pd.DataFrame, day_start: float = 6, day_end: float = 18) -> pd.DataFrame:
    result = frame.copy()
    hour_decimal = result["DateTime"].dt.hour + result["DateTime"].dt.minute / 60
    is_day = hour_decimal.between(day_start, day_end, inclusive="left")
    day_label = f"Day ({day_start:g}–{day_end:g})"
    night_label = f"Night ({day_end:g}–{day_start:g})"
    result["Period"] = np.where(is_day, day_label, night_label)
    result["PeriodDate"] = np.where(
        is_day,
        result["DateTime"].dt.date,
        (result["DateTime"] - pd.Timedelta(hours=day_start)).dt.date,
    )
    result["PeriodDate"] = pd.to_datetime(result["PeriodDate"])
    return result


def min_avg_max(frame: pd.DataFrame, group_columns: list[str], value: str) -> pd.DataFrame:
    return (
        frame.groupby(group_columns, observed=True)[value]
        .agg(Min="min", Avg="mean", Max="max")
        .reset_index()
    )


# ============================================================================
# PART 1 — HUMIDITY
# ============================================================================
def run_humidity(
    humidity_csv: Path,
    temperature_xlsx: Path,
    output_dir: Path,
    rh_low: float = 40,
    rh_high: float = 75,
    day_start: float = 7,
    day_end: float = 19,
):
    output_dir.mkdir(parents=True, exist_ok=True)
    humidity = read_greenhouse_source(humidity_csv, HUMIDITY_NAMES, "Humidity CSV")
    if humidity.shape[1] != 10 or humidity.empty:
        raise ValueError(f"Expected 10 non-empty humidity columns; found shape {humidity.shape}.")
    humidity[["Time", "Date"]] = humidity[["Time", "Date"]].apply(lambda x: x.str.strip())
    sensor_columns = HUMIDITY_NAMES[2:]
    humidity[sensor_columns] = humidity[sensor_columns].apply(pd.to_numeric, errors="coerce")
    if humidity[sensor_columns].isna().any().any():
        warnings.warn("One or more humidity readings are missing or nonnumeric.")

    temp_time = read_temperature_source(temperature_xlsx)
    require_columns(temp_time, ["Time", "Date"], "Temperature workbook")
    temp_time = temp_time[["Time", "Date"]].apply(lambda x: x.str.strip())
    if len(humidity) != len(temp_time):
        raise ValueError("Humidity and temperature row counts differ; timestamps cannot be matched safely.")
    mismatches = np.flatnonzero(humidity["Date"].to_numpy() != temp_time["Date"].to_numpy())
    if len(mismatches):
        raise ValueError(f"Humidity and temperature dates differ in data row(s): {(mismatches[:20] + 1).tolist()}")

    humidity_dt = parse_datetime(humidity["Date"], humidity["Time"], "Humidity data")
    temperature_dt = parse_datetime(temp_time["Date"], temp_time["Time"], "Temperature data")
    difference = (humidity_dt - temperature_dt).dt.total_seconds().abs()
    if (difference > 2).any():
        bad = (np.flatnonzero(difference > 2)[:20] + 1).tolist()
        raise ValueError(f"Humidity and temperature times differ by over two seconds in data row(s): {bad}")
    adjusted = int((humidity["Time"].to_numpy() != temp_time["Time"].to_numpy()).sum())
    humidity["Time"] = temp_time["Time"]

    negative_mask = humidity[sensor_columns] < 0
    negative_rows = []
    for row, column in zip(*np.where(negative_mask.to_numpy())):
        negative_rows.append({"Row": row + 1, "Date": humidity.at[row, "Date"], "Time": humidity.at[row, "Time"],
                              "Sensor": sensor_columns[column], "Value": humidity.at[row, sensor_columns[column]]})
    negative_report = pd.DataFrame(negative_rows)
    if not negative_report.empty:
        warnings.warn(f"{len(negative_report)} negative humidity reading(s) were retained.")
        write_single_sheet(negative_report, output_dir / "Negative humidity readings.xlsx", "Negative readings")

    formatted_path = output_dir / "ALLHUMI-20260206_formatted.xlsx"
    write_single_sheet(humidity, formatted_path, "Sheet 1 - ALLHUMI-20260206")
    workbook = load_workbook(formatted_path)
    sheet = workbook["Sheet 1 - ALLHUMI-20260206"]
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "@"
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=10):
        for cell in row:
            cell.number_format = "0"
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column_cells) + 2, 30
        )
    workbook.save(formatted_path)

    processed = humidity.rename(columns=lambda c: re.sub(r"^HUMI\s+", "", c))
    humi_sensors = [c for c in processed.columns if re.fullmatch(r"[ABCD]\d+", c)]
    if processed.isna().any().any():
        raise ValueError("NA detected in humidity data; fix the source before processing averages.")
    write_single_sheet(processed[["Time", "Date"] + humi_sensors], output_dir / "ALLHUMI processed.xlsx")

    long = processed.melt(["Time", "Date"], humi_sensors, "Sensor", "Humi")
    long["Dept"] = long["Sensor"].str[0]
    averages = long.groupby(["Time", "Date", "Dept"], as_index=False)["Humi"].mean()
    averages = averages.pivot(index=["Time", "Date"], columns="Dept", values="Humi").reset_index()
    averages.columns.name = None
    averages["DateTime"] = parse_datetime(averages["Date"], averages["Time"], "Humidity averages")
    averages = averages.sort_values("DateTime")
    export_avg = averages[["Time", "Date"] + DEPARTMENTS]
    write_single_sheet(export_avg, output_dir / "Humi averages processed.xlsx")
    write_single_sheet(export_avg, output_dir / "Humi averages.xlsx")

    humidity_long = averages.melt(["Time", "Date", "DateTime"], DEPARTMENTS, "Dept", "Humi")
    daily = min_avg_max(humidity_long.assign(Day=humidity_long.DateTime.dt.date), ["Day", "Dept"], "Humi")
    write_single_sheet(daily, output_dir / "Daily humidity statistics.xlsx")
    parts = add_period_columns(averages, day_start, day_end)
    period_long = parts.melt(["PeriodDate", "Period"], DEPARTMENTS, "Dept", "Humi")
    period_stats = min_avg_max(period_long, ["PeriodDate", "Period", "Dept"], "Humi")
    write_single_sheet(period_stats, output_dir / "Day and night humidity statistics.xlsx")

    print(f"Humidity complete: {len(humidity)} rows; {adjusted} times synchronized; output: {output_dir}")
    return output_dir / "Humi averages processed.xlsx", {
        "negative_readings": negative_report,
        "daily_humidity": daily,
        "daynight_humidity": period_stats,
    }


# ============================================================================
# PART 2 — TEMPERATURE
# ============================================================================
def run_temperature(
    temperature_xlsx: Path,
    output_dir: Path,
    scale: float = 10,
    day_low: float = 20,
    day_high: float = 35,
    night_low: float = 14,
    night_high: float = 18,
    day_start: float = 7,
    day_end: float = 19,
):
    output_dir.mkdir(parents=True, exist_ok=True)
    raw = read_temperature_source(temperature_xlsx)
    raw.columns = [re.sub(r"^TEMP\s+", "", str(c).strip()) for c in raw.columns]
    required = ["Time", "Date"] + TEMP_SENSORS
    require_columns(raw, required, "Temperature workbook")
    raw = raw[required].copy()
    raw[["Time", "Date"]] = raw[["Time", "Date"]].astype(str).apply(lambda x: x.str.strip())
    raw[TEMP_SENSORS] = raw[TEMP_SENSORS].apply(pd.to_numeric, errors="coerce") / scale
    if raw.isna().any().any():
        raise ValueError("NA values found after importing and converting temperature data.")
    raw["DateTime"] = parse_datetime(raw["Date"], raw["Time"], "Temperature data")
    raw = raw.sort_values("DateTime")
    write_single_sheet(raw[["Time", "Date"] + TEMP_SENSORS], output_dir / "ALLTEMP processed.xlsx")

    long = raw.melt(["Time", "Date", "DateTime"], TEMP_SENSORS, "Sensor", "Temperature")
    long["Dept"] = long["Sensor"].str[0]
    averages = long.groupby(["DateTime", "Dept"], as_index=False)["Temperature"].mean()
    wide = averages.pivot(index="DateTime", columns="Dept", values="Temperature").reset_index()
    wide.columns.name = None
    wide.insert(0, "Date", wide["DateTime"].dt.date)
    wide.insert(0, "Time", wide["DateTime"].dt.strftime("%H:%M:%S"))
    write_single_sheet(wide[["Time", "Date"] + DEPARTMENTS], output_dir / "Temp averages processed.xlsx")

    avg_long = wide.melt(["Time", "Date", "DateTime"], DEPARTMENTS, "Dept", "Temp")
    daily = min_avg_max(avg_long.assign(Day=avg_long.DateTime.dt.date), ["Day", "Dept"], "Temp")
    write_single_sheet(daily, output_dir / "Daily temperature statistics.xlsx")

    parts = add_period_columns(wide, day_start, day_end)
    part_long = parts.melt(["DateTime", "PeriodDate", "Period"], DEPARTMENTS, "Dept", "Temp")
    period_stats = min_avg_max(part_long, ["PeriodDate", "Period", "Dept"], "Temp")
    write_single_sheet(period_stats, output_dir / "Day and night temperature statistics.xlsx")

    breach_levels = [
        f"Night Low <{night_low:g}", f"Night High >{night_high:g}",
        f"Day Low <{day_low:g}", f"Day High >{day_high:g}",
    ]
    conditions = [
        (part_long.Period.str.startswith("Night") & part_long.Temp.lt(night_low)),
        (part_long.Period.str.startswith("Night") & part_long.Temp.gt(night_high)),
        (part_long.Period.str.startswith("Day") & part_long.Temp.lt(day_low)),
        (part_long.Period.str.startswith("Day") & part_long.Temp.gt(day_high)),
    ]
    part_long["Breach"] = np.select(conditions, breach_levels, default=None)
    breaches = part_long[part_long.Breach.notna()].copy()
    breaches["Date"] = breaches.DateTime.dt.date
    breaches["Time"] = breaches.DateTime.dt.strftime("%H:%M:%S")
    breaches["Temp"] = breaches.Temp.round(2)
    breaches = breaches[["Dept", "Breach", "Period", "PeriodDate", "Date", "Time", "DateTime", "Temp"]]
    counts = breaches.groupby(["PeriodDate", "Dept", "Breach"], observed=True).size().rename("Count").reset_index()
    wide_counts = counts.pivot_table(index=["PeriodDate", "Dept"], columns="Breach", values="Count", fill_value=0).reset_index()
    wide_counts.columns.name = None
    for level in breach_levels:
        if level not in wide_counts:
            wide_counts[level] = 0
    with pd.ExcelWriter(output_dir / "Temperature breaches.xlsx") as writer:
        breaches.to_excel(writer, sheet_name="Exact breaches", index=False)
        counts.to_excel(writer, sheet_name="Counts - long", index=False)
        wide_counts[["PeriodDate", "Dept"] + breach_levels].to_excel(
            writer, sheet_name="Counts - wide", index=False
        )

    print(f"Temperature complete: {len(raw)} rows; output: {output_dir}")
    return output_dir / "Temp averages processed.xlsx", {
        "daily_temperature": daily,
        "daynight_temperature": period_stats,
        "breaches": breaches,
        "breach_counts": wide_counts,
    }


# ============================================================================
# PART 3 — VPD, DEW POINT, GDH, SETPOINTS, AND CONDENSATION
# ============================================================================
def read_average_file(path: Path, label: str) -> pd.DataFrame:
    frame = pd.read_excel(path)
    require_columns(frame, ["Time", "Date"] + DEPARTMENTS, label)
    frame = frame[["Time", "Date"] + DEPARTMENTS].copy()
    frame["DateTime"] = parse_datetime(frame["Date"], frame["Time"], label)
    if frame.DateTime.duplicated().any():
        raise ValueError(f"{label} contains duplicate timestamps.")
    return frame


def run_vpd(
    temp_averages: Path,
    humi_averages: Path,
    output_dir: Path,
    cond_margin: float = 2,
    vpd_low: float = 0.8,
    vpd_high: float = 1.1,
    day_start: float = 7,
    day_end: float = 19,
    irrigation_start: float = 8.5,
    irrigation_end: float = 14.5,
    temp_day_low: float = 20,
    temp_day_high: float = 35,
    temp_night_low: float = 14,
    temp_night_high: float = 18,
    rh_low: float = 40,
    rh_high: float = 75,
):
    output_dir.mkdir(parents=True, exist_ok=True)
    temp = read_average_file(temp_averages, "Temperature averages file")
    humi = read_average_file(humi_averages, "Humidity averages file")
    temp_times, humi_times = set(temp.DateTime), set(humi.DateTime)
    if temp_times != humi_times:
        raise ValueError(
            f"Timestamps are not identical. Temperature-only: {len(temp_times - humi_times)}; "
            f"humidity-only: {len(humi_times - temp_times)}"
        )

    temp_long = temp.melt("DateTime", DEPARTMENTS, "Dept", "Temp")
    humi_long = humi.melt("DateTime", DEPARTMENTS, "Dept", "RH")
    climate = temp_long.merge(humi_long, on=["DateTime", "Dept"], validate="one_to_one").sort_values(["DateTime", "Dept"])
    invalid = climate.RH.lt(0) | climate.RH.gt(100)
    if invalid.any():
        warnings.warn(f"{int(invalid.sum())} invalid RH value(s) converted to NA for VPD calculations.")
        climate.loc[invalid, "RH"] = np.nan

    es = 0.6108 * np.exp((17.27 * climate.Temp) / (climate.Temp + 237.3))
    climate["VPD_kPa"] = es * (1 - climate.RH / 100)
    gamma = np.log(climate.RH / 100) + (17.27 * climate.Temp) / (237.7 + climate.Temp)
    climate["DewPoint_C"] = 237.7 * gamma / (17.27 - gamma)
    climate["T_minus_Dew_C"] = climate.Temp - climate.DewPoint_C
    climate["GDH_base10_step"] = np.maximum(climate.Temp - 10, 0) * (5 / 60)
    climate["VPD_zone"] = pd.cut(climate.VPD_kPa, [-np.inf, .4, .8, 1.1, 1.3, np.inf], right=False,
        labels=["Too humid (<0.4)", "Low/vegetative (0.4–0.8)", "Optimal (0.8–1.1)", "High/generative (1.1–1.3)", "Too dry (>1.3)"])
    climate["Date"] = climate.DateTime.dt.date
    climate["HourDec"] = climate.DateTime.dt.hour + climate.DateTime.dt.minute / 60
    climate["IsDay"] = climate.HourDec.between(day_start, day_end, inclusive="left")
    climate["IsIrrigationWindow"] = climate.HourDec.between(irrigation_start, irrigation_end, inclusive="both")
    climate["DayNight"] = np.where(climate.IsDay, "Day", "Night")
    climate["IrrigationFlag"] = np.where(climate.IsIrrigationWindow, "Irrigation", "Non-irrigation")
    climate["DayNight_6to6"] = np.where(climate.IsDay, "Day", "Night")
    climate["TempSetpointLow"] = np.where(climate.IsDay, temp_day_low, temp_night_low)
    climate["TempSetpointHigh"] = np.where(climate.IsDay, temp_day_high, temp_night_high)
    climate["RHSetpointLow"], climate["RHSetpointHigh"] = rh_low, rh_high
    climate["Temp_below_set"] = climate.Temp < climate.TempSetpointLow
    climate["Temp_above_set"] = climate.Temp > climate.TempSetpointHigh
    climate["RH_below_set"] = climate.RH < climate.RHSetpointLow
    climate["RH_above_set"] = climate.RH > climate.RHSetpointHigh
    climate["CondensationRisk"] = climate.T_minus_Dew_C < cond_margin

    grouped = climate.groupby(["Date", "Dept"], observed=True)
    daily = grouped.agg(Temp_mean=("Temp", "mean"), Temp_min=("Temp", "min"), Temp_max=("Temp", "max"),
        RH_mean=("RH", "mean"), RH_min=("RH", "min"), RH_max=("RH", "max"),
        VPD_mean=("VPD_kPa", "mean"), VPD_min=("VPD_kPa", "min"), VPD_max=("VPD_kPa", "max"),
        GDH_base10_sum=("GDH_base10_step", "sum"), pct_Temp_below_set=("Temp_below_set", "mean"),
        pct_Temp_above_set=("Temp_above_set", "mean"), pct_RH_below_set=("RH_below_set", "mean"),
        pct_RH_above_set=("RH_above_set", "mean"), pct_CondensationRisk=("CondensationRisk", "mean")).reset_index()
    pct_columns = [c for c in daily if c.startswith("pct_")]
    daily[pct_columns] *= 100
    zones = climate.dropna(subset=["VPD_zone"]).groupby(["Date", "Dept", "VPD_zone"], observed=True).size().rename("n").reset_index()
    zones["pct"] = zones.n / zones.groupby(["Date", "Dept"], observed=True).n.transform("sum") * 100
    flag_columns = ["Temp_below_set", "Temp_above_set", "RH_below_set", "RH_above_set", "CondensationRisk"]
    out = climate[climate[flag_columns].any(axis=1)].copy()
    out.insert(0, "Time_only", out.DateTime.dt.strftime("%H:%M:%S"))
    out.insert(0, "Date_only", out.DateTime.dt.date)
    workbook_path = output_dir / "climate_out_of_setpoints_and_condensation.xlsx"
    with pd.ExcelWriter(workbook_path) as writer:
        out.to_excel(writer, sheet_name="Out of setpoints", index=False)
        daily.to_excel(writer, sheet_name="Daily summary", index=False)
        zones.to_excel(writer, sheet_name="VPD zone daily", index=False)
        climate.to_excel(writer, sheet_name="Full climate data", index=False)

    print(f"VPD analysis complete: {len(climate)} department/readings; output: {output_dir}")
    return workbook_path, {
        "daily_summary": daily,
        "vpd_zones": zones,
        "out_of_setpoints": out,
        "climate": climate,
    }


# ============================================================================
# COMPLETE PIPELINE — runs everything, returns JSON-ready tables
# ============================================================================
def run_complete_analysis(
    temperature_file: Path,
    humidity_file: Path,
    output_dir: Path,
    temperature_scale: float = 10,
    condensation_margin: float = 2,
    vpd_low: float = 0.8,
    vpd_high: float = 1.1,
    temp_day_low: float = 20,
    temp_day_high: float = 35,
    temp_night_low: float = 14,
    temp_night_high: float = 18,
    rh_low: float = 40,
    rh_high: float = 75,
    day_start: float = 7,
    day_end: float = 19,
    irrigation_start: float = 8.5,
    irrigation_end: float = 14.5,
) -> dict:
    """Run the whole pipeline and return a dict of tables + paths (JSON-ready)."""
    output_dir.mkdir(parents=True, exist_ok=True)

    humi_avg, humi_tables = run_humidity(
        humidity_file, temperature_file, output_dir / "humidity",
        rh_low, rh_high, day_start, day_end,
    )
    temp_avg, temp_tables = run_temperature(
        temperature_file, output_dir / "temperature", temperature_scale,
        temp_day_low, temp_day_high, temp_night_low, temp_night_high, day_start, day_end,
    )
    vpd_book, vpd_tables = run_vpd(
        temp_avg, humi_avg, output_dir / "vpd", condensation_margin,
        vpd_low, vpd_high, day_start, day_end, irrigation_start, irrigation_end,
        temp_day_low, temp_day_high, temp_night_low, temp_night_high, rh_low, rh_high,
    )

    tables = {}
    for name, frame in {**humi_tables, **temp_tables, **vpd_tables}.items():
        tables[name] = _frame_to_records(frame)

    # Overview metrics
    daily = vpd_tables["daily_summary"]
    bc = temp_tables["breach_counts"]
    breach_columns = [c for c in bc.columns if c not in ("PeriodDate", "Dept")]
    metrics = {
        "days": int(daily["Date"].nunique()),
        "departments": int(daily["Dept"].nunique()),
        "temp_mean": float(daily["Temp_mean"].mean()),
        "vpd_mean": float(daily["VPD_mean"].mean()),
        "rh_mean": float(daily["RH_mean"].mean()),
        "gdh_total": float(daily["GDH_base10_sum"].sum()),
        "condensation_days": int((daily["pct_CondensationRisk"] > 0).sum()),
        "breach_counts": {level: int(bc[level].sum()) for level in breach_columns},
        "settings": {
            "temperature_scale": temperature_scale,
            "condensation_margin": condensation_margin,
            "vpd_low": vpd_low,
            "vpd_high": vpd_high,
            "temp_day_low": temp_day_low,
            "temp_day_high": temp_day_high,
            "temp_night_low": temp_night_low,
            "temp_night_high": temp_night_high,
            "rh_low": rh_low,
            "rh_high": rh_high,
            "day_start": day_start,
            "day_end": day_end,
            "irrigation_start": irrigation_start,
            "irrigation_end": irrigation_end,
        },
    }

    return {
        "tables": tables,
        "metrics": metrics,
        "output_dir": str(output_dir),
        "vpd_workbook": str(vpd_book),
    }


def _json_safe(v):
    """Robustly convert any pandas/numpy/python value into a JSON-serializable one."""
    if v is None or v is pd.NA:
        return None
    try:
        if isinstance(v, float) and (v != v):  # NaN
            return None
    except Exception:
        pass
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, np.bool_):
        return bool(v)
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        return str(v)
    if isinstance(v, (np.ndarray, list, tuple)):
        return [_json_safe(x) for x in v]
    if isinstance(v, (dict,)):
        return {str(k): _json_safe(x) for k, x in v.items()}
    return v


def _frame_to_records(frame: pd.DataFrame) -> list[dict]:
    """Convert a DataFrame to JSON-safe records (dates -> ISO strings, np types -> python)."""
    if frame is None or frame.empty:
        return []
    out = frame.copy()
    for col in out.columns:
        if isinstance(out[col].dtype, pd.CategoricalDtype):
            out[col] = out[col].astype("object")
    records = out.to_dict(orient="records")
    cleaned = []
    for rec in records:
        item = {}
        for k, v in rec.items():
            item[str(k)] = _json_safe(v)
        cleaned.append(item)
    return cleaned
