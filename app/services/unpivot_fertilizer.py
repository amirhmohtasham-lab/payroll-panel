"""Fertilizer Excel unpivot: wide sheet -> long (dept, crop, amount) rows."""

import sys, os, re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

try:
    import jdatetime
except ImportError:
    jdatetime = None

# ─── CONFIG ──────────────────────────────────────────────────────────────────

INPUT_COLUMNS = {
    'rec_date':1,'stage':2,'exec_date':3,'gap':4,'rec_num':5,
    'well':6,'area':7,'plant':8,'variety1':9,'variety2':10,'variety3':11,
}
FERT_NAME_START=12; REC_HA_START=22; REC_AREA_START=32
CONS_WEIGHT_START=42; INVENTORY_START=52; NUM_SLOTS=10
DV_SHEET='Data Validation'
DV_COL_NAME=1; DV_COL_TYPE=2; DV_COL_UNIT=3; DV_COL_PRICE=4; DV_COL_INIT_INV=5

TARGET_COLUMNS = [
    'ردیف','تاریخ توصیه','شماره سرک','تاریخ اجرا','فاصله (روز)',
    'شماره توصیه','شماره چاه','مساحت (هکتار)','نوع گیاه',
    'واریته ۱','واریته ۲','واریته ۳',
    'نام کود','قیمت فی خرید','توصیه/هکتار','توصیه/مساحت',
    'مصرفی/وزنی','مصرفی/ریالی','موجودی انبار','واحد','جنس کود',
    'مازاد/کمبود','تحقق%','پرچم',
]

PERSIAN_MONTH_DAYS = {1:31,2:31,3:31,4:31,5:31,6:31,7:30,8:30,9:30,10:30,11:30,12:29}
_ARABIC_CHARS = str.maketrans({'ي':'ی','ى':'ی','ئ':'ی','ك':'ک','ة':'ه','ۀ':'ه','أ':'ا','إ':'ا','آ':'آ','ؤ':'و','ٶ':'و','۰':'0','۱':'1','۲':'2','۳':'3','۴':'4','۵':'5','۶':'6','۷':'7','۸':'8','۹':'9'})

GREEN_FILL=PatternFill(start_color='C6EFCE',end_color='C6EFCE',fill_type='solid')
YELLOW_FILL=PatternFill(start_color='FFEB9C',end_color='FFEB9C',fill_type='solid')
RED_FILL=PatternFill(start_color='FFC7CE',end_color='FFC7CE',fill_type='solid')


def safe_float(v,d=0.0):
    if v is None: return d
    try: return float(v)
    except: return d

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

def normalize_fert_name(name):
    if name is None: return ''
    name=str(name).strip().translate(_ARABIC_CHARS)
    name=re.sub(r'\s+',' ',name)
    return name.strip(' .-،,')

def validate_jalali(ds):
    if not ds or ds=='-': return True,ds,None
    ds=str(ds).strip()
    m=re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$',ds)
    if not m: return False,ds,f"Bad format: {ds}"
    y,m_,d_=int(m.group(1)),int(m.group(2)),int(m.group(3))
    if m_<1 or m_>12: return False,ds,f"Bad month: {m_}"
    md=PERSIAN_MONTH_DAYS.get(m_,30)
    if d_<1 or d_>md: return False,ds,f"Bad day: {d_}"
    return True,f"{y:04d}/{m_:02d}/{d_:02d}",None

def compute_gap(a_str,b_str):
    if not a_str or a_str=='-' or not b_str or b_str=='-': return None
    if jdatetime is None: return None
    try:
        a=[int(x) for x in str(a_str).strip().split('/')]
        b=[int(x) for x in str(b_str).strip().split('/')]
        if len(a)!=3 or len(b)!=3: return None
        return (jdatetime.date(*b)-jdatetime.date(*a)).days
    except: return None

def build_fertilizer_lookup(ws_dv):
    lookup={}
    for r in range(2,ws_dv.max_row+1):
        raw=ws_dv.cell(row=r,column=DV_COL_NAME).value
        if raw and str(raw).strip():
            nn=normalize_fert_name(raw)
            if nn not in lookup:
                lookup[nn]={'price':safe_float(ws_dv.cell(row=r,column=DV_COL_PRICE).value),'unit':safe_str(ws_dv.cell(row=r,column=DV_COL_UNIT).value),'type':safe_str(ws_dv.cell(row=r,column=DV_COL_TYPE).value),'init_inv':safe_float(ws_dv.cell(row=r,column=DV_COL_INIT_INV).value)}
    return lookup

def compute_rec_per_area(rph,area,cached):
    if cached and cached>0: return cached
    return rph*area

def compute_flags(rph,rpa,cw,ach,inv):
    f=[]
    if rph<=0 and cw>0: f.append('مصرف بدون توصیه')
    if rph>0 and cw==0: f.append('توصیه انجام نشده')
    if 0<ach<50: f.append(f'تحقق پایین ({ach}%)')
    if ach>150: f.append(f'تحقق بالا ({ach}%)')
    if inv<0: f.append('موجودی منفی')
    return ' | '.join(f) if f else ''

def unpivot_data(ws_input,ws_input_data,fert_lookup):
    raw=[]
    for src_row in range(3,ws_input.max_row+1):
        rd_raw=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['rec_date']).value
        if rd_raw is None: continue
        rec_date=safe_str(rd_raw)
        v,rec_date,_=validate_jalali(rec_date)
        stage_num=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['stage']).value
        ed_raw=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['exec_date']).value
        exec_date=safe_str(ed_raw)
        v2,exec_date,_=validate_jalali(exec_date)
        gap_days=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['gap']).value
        rec_num=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['rec_num']).value
        well_num=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['well']).value
        area=safe_float(ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['area']).value)
        plant_type=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['plant']).value
        var1=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['variety1']).value
        var2=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['variety2']).value
        var3=ws_input_data.cell(row=src_row,column=INPUT_COLUMNS['variety3']).value
        if isinstance(stage_num,(int,float)): stage_num=str(int(stage_num))
        else: stage_num=safe_str(stage_num)
        gap_val=compute_gap(rec_date,exec_date)
        if gap_val is None:
            try: gap_val=int(float(gap_days)) if gap_days is not None else 0
            except: gap_val=0
        for slot in range(NUM_SLOTS):
            fert_raw=ws_input.cell(row=src_row,column=FERT_NAME_START+slot).value
            if fert_raw is None or str(fert_raw).strip()=='': continue
            fert_name=normalize_fert_name(fert_raw)
            if not fert_name: continue
            rph=safe_float(ws_input_data.cell(row=src_row,column=REC_HA_START+slot).value)
            rac=safe_float(ws_input_data.cell(row=src_row,column=REC_AREA_START+slot).value)
            rpa=compute_rec_per_area(rph,area,rac)
            cw=safe_float(ws_input_data.cell(row=src_row,column=CONS_WEIGHT_START+slot).value)
            inv_raw=safe_float(ws_input_data.cell(row=src_row,column=INVENTORY_START+slot).value)
            finfo=fert_lookup.get(fert_name,{})
            price=finfo.get('price',0.0)
            unit=finfo.get('unit','')
            ftype=finfo.get('type','')
            init_inv=finfo.get('init_inv',0.0)
            cons_rial=cw*price
            surplus=round(rpa-cw,2)
            ach=round((cw/rpa)*100,2) if rpa>0 else 0.0
            flag=compute_flags(rph,rpa,cw,ach,0)
            raw.append({'fert_name':fert_name,'exec_date':exec_date,'rec_date':rec_date,'cons_weight':cw,'init_inv':init_inv,'inventory':0.0,'flag':flag,'achievement':ach,'rec_per_area':rpa,'price':price,'unit':unit,'ftype':ftype,'row_data':[None,rec_date,stage_num,exec_date,gap_val,safe_str(rec_num),well_num,area,plant_type,var1,var2,var3,fert_name,round(price,0),rph,rpa,cw,round(cons_rial,0),None,unit,ftype,surplus,ach,None]})
    return raw

def compute_running_inventory(raw):
    fert_groups=defaultdict(list)
    for idx,rec in enumerate(raw): fert_groups[rec['fert_name']].append((idx,rec))
    for group in fert_groups.values():
        group.sort(key=lambda x:x[1]['exec_date'])
        running=group[0][1]['init_inv']
        for orig_idx,rec in group:
            running-=rec['cons_weight']
            raw[orig_idx]['inventory']=round(running,2)
    for rec in raw:
        rec['flag']=compute_flags(rec['row_data'][14],rec['rec_per_area'],rec['cons_weight'],rec['achievement'],rec['inventory'])

def write_cleaned_sheet(wb,cleaned):
    if 'Cleaned Data' in wb.sheetnames: del wb['Cleaned Data']
    ws=wb.create_sheet('Cleaned Data')
    hf=Font(name='B Nazanin',bold=True,size=11,color='FFFFFF')
    hfill=PatternFill(start_color='063B5B',end_color='063B5B',fill_type='solid')
    ha=Alignment(horizontal='center',vertical='center',wrap_text=True)
    df=Font(name='B Nazanin',size=10)
    da=Alignment(horizontal='center',vertical='center')
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for ci,cn in enumerate(TARGET_COLUMNS,1):
        c=ws.cell(row=1,column=ci,value=cn); c.font=hf; c.fill=hfill; c.alignment=ha; c.border=tb
    for ri,row_data in enumerate(cleaned,2):
        for ci,val in enumerate(row_data,1):
            c=ws.cell(row=ri,column=ci,value=val); c.font=df; c.alignment=da; c.border=tb
            if isinstance(val,float):
                if abs(val)>=1000000: c.number_format='#,##0'
                elif val==int(val): c.number_format='#,##0'
                else: c.number_format='#,##0.00'
            elif isinstance(val,int): c.number_format='#,##0'
    ldr=len(cleaned)+1
    ws.conditional_formatting.add(f"W2:W{ldr}",CellIsRule(operator='greaterThan',formula=['150'],fill=RED_FILL,font=Font(color='9C0006')))
    ws.conditional_formatting.add(f"W2:W{ldr}",CellIsRule(operator='lessThan',formula=['50'],fill=RED_FILL,font=Font(color='9C0006')))
    ws.conditional_formatting.add(f"W2:W{ldr}",CellIsRule(operator='between',formula=['50','80'],fill=YELLOW_FILL,font=Font(color='9C6500')))
    ws.conditional_formatting.add(f"W2:W{ldr}",CellIsRule(operator='between',formula=['120','150'],fill=YELLOW_FILL,font=Font(color='9C6500')))
    ws.conditional_formatting.add(f"W2:W{ldr}",CellIsRule(operator='between',formula=['80','120'],fill=GREEN_FILL,font=Font(color='006100')))
    ws.conditional_formatting.add(f"S2:S{ldr}",CellIsRule(operator='lessThan',formula=['0'],fill=RED_FILL,font=Font(color='9C0006')))
    ws.conditional_formatting.add(f"V2:V{ldr}",CellIsRule(operator='lessThan',formula=['0'],fill=RED_FILL,font=Font(color='9C0006')))
    ws.conditional_formatting.add(f"V2:V{ldr}",CellIsRule(operator='greaterThan',formula=['0'],fill=GREEN_FILL,font=Font(color='006100')))
    ws.conditional_formatting.add(f"X2:X{ldr}",FormulaRule(formula=['X2<>""'],fill=YELLOW_FILL))
    widths={1:7,2:14,3:12,4:14,5:10,6:12,7:10,8:12,9:14,10:14,11:14,12:14,13:20,14:16,15:14,16:14,17:14,18:18,19:16,20:10,21:10,22:14,23:10,24:30}
    for ci,w in widths.items(): ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes='A2'
    ws.auto_filter.ref=f"A1:X{len(cleaned)+1}"
    return len(cleaned)

def write_summary_sheet(wb,raw,fert_lookup):
    if 'Summary' in wb.sheetnames: del wb['Summary']
    ws=wb.create_sheet('Summary')
    well_data=defaultdict(lambda:{'count':0,'weight':0.0,'rial':0.0})
    fert_data=defaultdict(lambda:{'count':0,'weight':0.0,'rial':0.0})
    stage_data=defaultdict(lambda:{'count':0,'ach_sum':0.0,'rec_area_sum':0.0,'cons_sum':0.0})
    fert_inv={}
    for rec in raw:
        w=rec['cons_weight'];r=rec['row_data'][17] or 0;ach=rec['achievement'];ra=rec['rec_per_area']
        well_data[str(rec['row_data'][6] or 'نامشخص')]['count']+=1; well_data[str(rec['row_data'][6] or 'نامشخص')]['weight']+=w; well_data[str(rec['row_data'][6] or 'نامشخص')]['rial']+=r
        fert_data[rec['fert_name']]['count']+=1; fert_data[rec['fert_name']]['weight']+=w; fert_data[rec['fert_name']]['rial']+=r
        st=str(rec['row_data'][2] or 'نامشخص'); stage_data[st]['count']+=1; stage_data[st]['ach_sum']+=ach; stage_data[st]['cons_sum']+=w; stage_data[st]['rec_area_sum']+=ra
        fert_inv[rec['fert_name']]=rec['inventory']
    tf=Font(name='B Nazanin',bold=True,size=13,color='063B5B')
    thf=Font(name='B Nazanin',bold=True,size=10,color='FFFFFF')
    thfill=PatternFill(start_color='063B5B',end_color='063B5B',fill_type='solid')
    df=Font(name='B Nazanin',size=10); da=Alignment(horizontal='center',vertical='center')
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    cr=1
    def wt(title,headers,rows,sr):
        r=sr; ws.cell(row=r,column=1,value=title).font=tf; r+=1
        for ci,h in enumerate(headers,1):
            c=ws.cell(row=r,column=ci,value=h); c.font=thf; c.fill=thfill; c.alignment=da; c.border=tb
        r+=1
        for row in rows:
            for ci,val in enumerate(row,1):
                c=ws.cell(row=r,column=ci,value=val); c.font=df; c.alignment=da; c.border=tb
                if isinstance(val,float):
                    if abs(val)>=1000000: c.number_format='#,##0'
                    elif val==int(val): c.number_format='#,##0'
                    else: c.number_format='#,##0.00'
            r+=1
        return r+1
    sw=sorted(well_data.items(),key=lambda x:x[1]['rial'],reverse=True)
    t1=[[w,d['count'],round(d['weight'],2),round(d['rial'],0)] for w,d in sw]
    cr=wt('📊 جدول ۱ — مصرف به تفکیک شماره چاه',['شماره چاه','تعداد','مصرف کل (وزنی)','مصرف کل (ریالی)'],t1,cr)
    sf=sorted(fert_data.items(),key=lambda x:x[1]['rial'],reverse=True)
    t2=[[f,finfo.get('unit',''),finfo.get('type',''),d['count'],round(d['weight'],2),round(d['rial'],0),round(finfo.get('init_inv',0),0),round(fert_inv.get(f,0),2),round((d['weight']/finfo.get('init_inv',1)*100),1) if finfo.get('init_inv',0)>0 else 0] for f,d in sf for finfo in [fert_lookup.get(f,{})]]
    cr=wt('📊 جدول ۲ — مصرف به تفکیک کود',['نام کود','واحد','جنس','تعداد','مصرف کل (وزنی)','مصرف کل (ریالی)','موجودی اولیه','موجودی فعلی','درصد مصرف'],t2,cr)
    def sk(item):
        s=item[0]
        if s=='قبل کشت': return (0,0)
        if s=='۱ و ۲': return (2,0)
        if s=='ویژه': return (3,0)
        try: return (1,int(s))
        except: return (4,s)
    ss=sorted(stage_data.items(),key=sk)
    t3=[[st,d['count'],round(d['ach_sum']/d['count'],1) if d['count']>0 else 0,round(d['rec_area_sum'],2),round(d['cons_sum'],2)] for st,d in ss]
    cr=wt('📊 جدول ۳ — تحقق به تفکیک شماره سرک',['شماره سرک','تعداد','میانگین تحقق%','مجموع توصیه/مساحت','مجموع مصرف'],t3,cr)
    flagged=[rec for rec in raw if rec['flag']]
    flagged.sort(key=lambda x:x['flag'])
    t4=[[rd[1] or '',rd[2] or '',rd[6] or '',rd[12] or '',rd[16] or 0,rd[17] or 0,rec['achievement'],rec['flag']] for rec in flagged[:50] for rd in [rec['row_data']]]
    if t4: cr=wt('⚠️ جدول ۴ — رکوردهای نیازمند توجه',['تاریخ اجرا','شماره سرک','شماره چاه','نام کود','مصرفی/وزنی','مصرفی/ریالی','تحقق%','توضیح'],t4,cr)
    for row in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value,str) and 'میانگین تحقق%' in str(cell.value):
                hr=cell.row; ds=hr+1; de=ds+len(t3)-1; ac=cell.column
                if de>=ds: rng=f"{get_column_letter(ac)}{ds}:{get_column_letter(ac)}{de}"; ws.conditional_formatting.add(rng,CellIsRule(operator='greaterThan',formula=['150'],fill=RED_FILL,font=Font(color='9C0006'))); ws.conditional_formatting.add(rng,CellIsRule(operator='lessThan',formula=['50'],fill=RED_FILL,font=Font(color='9C0006'))); ws.conditional_formatting.add(rng,CellIsRule(operator='between',formula=['80','120'],fill=GREEN_FILL,font=Font(color='006100')))
                break
    for row in ws.iter_rows(min_row=1,max_row=ws.max_row,max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value,str) and 'درصد مصرف' in str(cell.value):
                hr=cell.row; ds=hr+1; de=ds+len(t2)-1; pc=cell.column
                if de>=ds: rng=f"{get_column_letter(pc)}{ds}:{get_column_letter(pc)}{de}"; ws.conditional_formatting.add(rng,CellIsRule(operator='greaterThan',formula=['100'],fill=RED_FILL,font=Font(color='9C0006')))
                break
    for ci in range(1,11): ws.column_dimensions[get_column_letter(ci)].width=20
    return True

# ─── RUN UNPIVOT (main entry for module) ────────────────────────────────────

def run_unpivot(file_path):
    """Run the full unpivot pipeline. Returns dict: {success, row_count, error}."""
    try:
        wb=load_workbook(file_path); wb_data=load_workbook(file_path,data_only=True)
        required={'اطلاعات ورودی',DV_SHEET}; missing=required-set(wb.sheetnames)
        if missing:
            return {"success":False,"row_count":0,"error":f"Missing sheets: {', '.join(missing)}"}
        fert_lookup=build_fertilizer_lookup(wb_data[DV_SHEET])
        raw=unpivot_data(wb['اطلاعات ورودی'],wb_data['اطلاعات ورودی'],fert_lookup)
        if not raw:
            return {"success":False,"row_count":0,"error":"No data rows found"}
        compute_running_inventory(raw)
        cleaned=[]
        for idx,rec in enumerate(raw):
            rec['row_data'][0]=idx+1; rec['row_data'][18]=rec['inventory']; rec['row_data'][23]=rec['flag']; cleaned.append(rec['row_data'])
        rc=write_cleaned_sheet(wb,cleaned)
        write_summary_sheet(wb,raw,fert_lookup)
        wb.save(file_path)
        return {"success":True,"row_count":rc,"error":None}
    except Exception as e:
        return {"success":False,"row_count":0,"error":str(e)}

def main():
    if len(sys.argv)<2: print("Usage: python3 unpivot_fertilizer_data.py <path>"); sys.exit(1)
    fp=sys.argv[1].strip()
    if not os.path.isfile(fp): print(f"❌ Not found: {fp}"); sys.exit(1)
    result=run_unpivot(fp)
    if result['success']:
        flagged_count=0
        try:
            wb=load_workbook(fp); ws=wb['Cleaned Data']
            for row in ws.iter_rows(min_row=2,max_row=ws.max_row,values_only=True):
                if len(row)>=24 and row[23]: flagged_count+=1
        except: pass
        print(f"\n{'='*60}\n✅ DONE!\n   📄 Cleaned Data: {result['row_count']} rows\n   📊 Summary: 4 aggregate tables\n   🚩 Flagged: {flagged_count}\n{'='*60}")
    else:
        print(f"❌ Error: {result['error']}"); sys.exit(1)

if __name__=='__main__': main()
